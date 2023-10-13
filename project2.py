import zipfile
import sys
import os
import io
import glob
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import BaggingClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import RandomizedSearchCV
from sklearn.metrics import make_scorer, accuracy_score, precision_score, recall_score, f1_score
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.model_selection import train_test_split
from sklearn.datasets import fetch_openml
class dataExtractor:
    def __init__(self, zipPath):
        self.allData = self._initializeAllData()
        self.zipPath = os.path.join("./", zipPath)
        self.folderName = None
        self._extractFile()
        self.bestParams = self._initialization()
        self.testResults = self._initialization()
        self.scoring = {
            'f1_score': make_scorer(f1_score, zero_division=0.0, average="binary"),
            'accuracy': make_scorer(accuracy_score),
            'precision': make_scorer(precision_score, zero_division=0.0, average="binary"),
            'recall': make_scorer(recall_score, average="binary")
        }

    
    def _initializeAllData(self):
        result = {}
        for clause in [300, 500, 1000, 1500, 1800]:
            clauseDict = {}
            for example in [100, 1000, 5000]:
                clauseDict[example] = clauseDict.get(example, {"train": None, "valid": None, "test": None})
            result[clause] = result.get(clause, clauseDict)
        return result
    def _extractFile(self):
        with zipfile.ZipFile(self.zipPath, "r") as folder:
            csvList = [_ for _ in folder.namelist() if _.endswith(".csv")]
            for oneCsv in csvList:
                csvName = oneCsv.split("/")[-1]
                dataType = csvName.split("_")[0]
                clauseNumber = int(csvName.split("_")[1][1:])
                exampleNumber = int(csvName.split("_")[2][1:].split(".")[0])
                with folder.open(oneCsv) as f:
                    dataSet = np.genfromtxt(f, delimiter=",", dtype=int)
                    self.allData[clauseNumber][exampleNumber][dataType] = dataSet
    def _initialization(self):
        result = {}
        for clause in [300, 500, 1000, 1500, 1800]:
            clauseDict = {}
            for example in [100, 1000, 5000]:
                clauseDict[example] = clauseDict.get(example, {"bestParams": None, "accuracy": None, "F1Score": None})
            result[clause] = result.get(clause, clauseDict)
        return result
    
    def _randomizedGridSearch(self, X, Y, model, paramGrid):
        model = model
        search = RandomizedSearchCV(estimator=model, param_distributions=paramGrid, n_iter=50, scoring=self.scoring, n_jobs=-2, refit="f1_score")
        search.fit(X, Y)
        return search.best_params_, search.best_estimator_
    def searchForEachDataSet(self, paramGrid, identifier):
        for clauseNumber, val1 in self.allData.items():
            for exampleNumber, val2 in val1.items():
                X, Y = val2["valid"][:, :-1], val2["valid"][:, -1]
                searchModel = None
                if identifier == "D":
                    searchModel = DecisionTreeClassifier()
                elif identifier == "BG":
                    searchModel = BaggingClassifier(DecisionTreeClassifier())
                elif identifier == "RF":
                    searchModel = RandomForestClassifier()
                elif identifier == "GB":
                    searchModel = GradientBoostingClassifier()
                self.bestParams[clauseNumber][exampleNumber]["bestParams"], model = self._randomizedGridSearch(X, Y, searchModel, paramGrid)
                self.testResults[clauseNumber][exampleNumber]["bestParams"] = self.bestParams[clauseNumber][exampleNumber]["bestParams"]
                [self.bestParams[clauseNumber][exampleNumber]["accuracy"], self.bestParams[clauseNumber][exampleNumber]["F1Score"]] = self._calc(Y, model.predict(X))
    
    def test(self, identifier):
        for clauseNumber, val1 in self.allData.items():
            for exampleNumber, val2 in val1.items():
                X, Y = np.vstack((self.allData[clauseNumber][exampleNumber]["train"][:,:-1], self.allData[clauseNumber][exampleNumber]["valid"][:,:-1])), np.concatenate((self.allData[clauseNumber][exampleNumber]["train"][:,-1], self.allData[clauseNumber][exampleNumber]["valid"][:,-1]))
                model = None
                if identifier == "D":
                    model = DecisionTreeClassifier(**(self.testResults[clauseNumber][exampleNumber]["bestParams"]))
                elif identifier == "BG":
                    model = BaggingClassifier(DecisionTreeClassifier(), **(self.testResults[clauseNumber][exampleNumber]["bestParams"]))
                elif identifier == "RF":
                    model = RandomForestClassifier(**(self.testResults[clauseNumber][exampleNumber]["bestParams"]))
                elif identifier == "GB":
                    model = GradientBoostingClassifier(**(self.testResults[clauseNumber][exampleNumber]["bestParams"]))
                model.fit(X, Y)
                YPred = model.predict(self.allData[clauseNumber][exampleNumber]["test"][:, :-1])
                [self.testResults[clauseNumber][exampleNumber]["accuracy"], self.testResults[clauseNumber][exampleNumber]["F1Score"]] = self._calc(self.allData[clauseNumber][exampleNumber]["test"][:, -1], YPred)
    
    def _calc(self, Y, YPred):
        accuracy = accuracy_score(Y, YPred)
        fscore = f1_score(Y, YPred, zero_division=0.0, average="binary")
        temp = [accuracy, fscore]
        return [round(_, 3) for _ in temp]
    
    def outputToTextFile(self, identifier):
        filePath = f"./results_{identifier}.txt"
        with open(filePath, "w") as file:
            sys.stdout = file
            for clauseNumber, val1 in self.allData.items():
                for exampleNumber, val2 in val1.items():
                    print(f"\nRandomized Grid Search Result, hyperparametes tuned on validation data, for the dataSet generated by {clauseNumber} clauses with {exampleNumber} examples: \n")
                    print("\nBest Params:\n")
                    print(self.bestParams[clauseNumber][exampleNumber]["bestParams"])
                    print("\nAccuracy\n")
                    print(self.bestParams[clauseNumber][exampleNumber]["accuracy"])
                    print("\nf1_score:\n")
                    print(self.bestParams[clauseNumber][exampleNumber]["F1Score"])
                    print("\n--------test accuracy and f1score\n")
                    print("\nAccuracy\n")
                    print(self.testResults[clauseNumber][exampleNumber]["accuracy"])
                    print("\nf1_score:\n")
                    print(self.testResults[clauseNumber][exampleNumber]["F1Score"])
                    print("\n--------test accuracy and f1score\n")

            sys.stdout = sys.__stdout__
    def exportTable(self, identifier):
        table = []
        for clause, val1 in self.testResults.items():
            for example, val2 in val1.items():
                row = {
                    "Clauses_number": clause,
                    "examples_number": example,
                    **{f"Parameter_{k}": v for k, v in val2["bestParams"].items()},
                    "validation_accuracy": self.bestParams[clause][example]["accuracy"],
                    "validation_f1score": self.bestParams[clause][example]["F1Score"],
                    "test_accuracy": val2["accuracy"],
                    "test_f1score": val2["F1Score"]
                }
                table.append(row)
        
        df = pd.DataFrame(table)
        xlsxFile = f"results_{identifier}.xlsx"
        df.to_excel(xlsxFile, index=False)
        file = load_workbook(xlsxFile)
        sheet = file.active
        sheet.insert_rows(idx=1)

        numColumns = len(df.columns)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=numColumns)
        sheet["A1"] = "D: decision tree classifier\nBG: bagging classifier\nRF: random forest classifier\nGB: gradient boosting classifier\nAuthor: Ian Zhang"
        file.save(xlsxFile)



class decisionTreeExperiment(dataExtractor):
    def __init__(self, zipPath):
        super().__init__(zipPath)
        self.paramGrid = {
            "criterion": ["gini", "entropy"],
            "splitter": ["best", "random"],
            "max_depth": list(range(2, 53, 2)),
            "max_features": [None, "sqrt", "log2"],
            "min_samples_split": list(range(2, 101, 2))
        }

class baggingClassifierExperiment(dataExtractor):
    def __init__(self, zipPath):
        super().__init__(zipPath)
        self.paramGrid = {
            "n_estimators": list(range(5, 101, 1)),
            "max_samples": np.linspace(0.1, 1.0, 10),
            "max_features": np.linspace(0.1, 1.0, 10)
        }

class randomForestClassifierExperiment(dataExtractor):
    def __init__(self, zipPath):
        super().__init__(zipPath)
        self.paramGrid = {
            "n_estimators": list(range(5, 101, 1)),
            "criterion": ["gini", "entropy"],
            "class_weight": ["balanced", "balanced_subsample"],
            "max_features": [None, 'sqrt', 'log2'],
            "max_depth":list(range(2, 53, 2)),
        }

class gradientBoostingClassifierExperiment(dataExtractor):
    def __init__(self, zipPath):
        super().__init__(zipPath)
        self.paramGrid = {
            "max_depth":list(range(2, 53, 2)),
            "n_estimators": list(range(100, 1100, 100)),
            "learning_rate": [0.1, 0.2, 0.3, 0.4, 0.5],
            "loss": ["log_loss"],
            "max_features": ["sqrt", "log2"]
        }

class mnistExperiment:
    def __init__(self):
        X, y = fetch_openml("mnist_784", version=1, return_X_y=True, parser="auto")
        X = X / 255.
        # rescale the data, use the traditional train/test split
        # (60K: Train) and (10K: Test)
        X_train, X_test = X[:60000], X[60000:]
        y_train, y_test = y[:60000], y[60000:]
        self.trainX = X_train
        self.trainY = y_train
        self.testX = X_test
        self.testY = y_test
        _, self.validationDataX, _y, self.validationDataY = train_test_split(self.trainX, self.trainY, test_size=0.1, stratify=y_train, random_state=4)
        self.scoring = {
            'f1_score': make_scorer(f1_score, zero_division=0.0, average="macro"),
            'accuracy': make_scorer(accuracy_score),
            'precision': make_scorer(precision_score, zero_division=0.0, average="macro"),
            'recall': make_scorer(recall_score, average="macro")
        }
        self.bastParams = None
        self.validationAccuracy = None
        self.testAccuracy = None
    def _randomizedGridSearch(self, model, paramGrid):
        model = model
        search = RandomizedSearchCV(estimator=model, param_distributions=paramGrid, n_iter=50, scoring=self.scoring, n_jobs=-2, refit="f1_score")
        search.fit(self.validationDataX, self.validationDataY)
        return search.best_params_, search.best_estimator_

    def search(self, paramGrid, identifier):
        searchModel = None
        if identifier == "D":
            searchModel = DecisionTreeClassifier()
        elif identifier == "BG":
            searchModel = BaggingClassifier(DecisionTreeClassifier())
        elif identifier == "RF":
            searchModel = RandomForestClassifier()
        elif identifier == "GB":
            searchModel = GradientBoostingClassifier()
        self.bestParams, model = self._randomizedGridSearch(searchModel, paramGrid)
        self.validationAccuracy = accuracy_score(self.validationDataY, model.predict(self.validationDataX))

    def test(self, identifier):
        model = None
        if identifier == "D":
            model = DecisionTreeClassifier(**(self.bestParams))
        elif identifier == "BG":
            model = BaggingClassifier(DecisionTreeClassifier(), **(self.bestParams))
        elif identifier == "RF":
            model = RandomForestClassifier(**(self.bestParams))
        elif identifier == "GB":
            model = GradientBoostingClassifier(**(self.bestParams))
        model.fit(self.trainX, self.trainY)
        self.testAccuracy = accuracy_score(self.testY, model.predict(self.testX))
    

def mnistResults():
    paramsDict = {"D": {
            "criterion": ["gini", "entropy"],
            "splitter": ["best", "random"],
            "max_depth": list(range(2, 53, 2)),
            "max_features": [None, "sqrt", "log2"],
            "min_samples_split": list(range(2, 101, 2))
        },
        "BG": {
            "n_estimators": list(range(5, 101, 1)),
            "max_samples": np.linspace(0.1, 1.0, 10),
            "max_features": np.linspace(0.1, 1.0, 10)
        },
        "RF": {
            "n_estimators": list(range(5, 101, 1)),
            "criterion": ["gini", "entropy"],
            "class_weight": ["balanced", "balanced_subsample"],
            "max_features": [None, 'sqrt', 'log2'],
            "max_depth":list(range(2, 53, 2)),
        },
        "GB": {
            "max_depth":list(range(2, 53, 2)),
            "n_estimators": list(range(100, 1100, 100)),
            "learning_rate": [0.1, 0.2, 0.3, 0.4, 0.5],
            "loss": ["log_loss"],
            "max_features": ["sqrt", "log2"]
        }}
    table = []
    for identifier, paramGrid in paramsDict.items():
        experiment = mnistExperiment()
        experiment.search(paramGrid, identifier)
        experiment.test(identifier)
        row = {
            "identifier": identifier,
            "bestParams": experiment.bastParams,
            "validationAccuracy": experiment.validationAccuracy,
            "testAccuracy": experiment.testAccuracy
        }
        table.append(row)
    
    df = pd.DataFrame(table)
    xlsxFile = "MNIST.xlsx"
    df.to_excel(xlsxFile, index=False)
    file = load_workbook(xlsxFile)
    sheet = file.active
    sheet.insert_rows(idx=1)

    numColumns = len(df.columns)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=numColumns)
    sheet["A1"] = "D: decision tree classifier\nBG: bagging classifier\nRF: random forest classifier\nGB: gradient boosting classifier\nAuthor: Ian Zhang"
    file.save(xlsxFile)
    
        
    
def experiment(zipPath):
    experiments = {"D": decisionTreeExperiment(zipPath), "BG": baggingClassifierExperiment(zipPath), "RF": randomForestClassifierExperiment(zipPath), "GB": gradientBoostingClassifierExperiment(zipPath)}
    for key, val in experiments.items():
        val.searchForEachDataSet(val.paramGrid, key)
        val.test(key)
        val.exportTable(key)
def main():
    experiment("project2_data.zip")
    mnistResults()


if __name__ == "__main__":
    main()


# ext = dataExtractor("project2_data.zip")